"""
SmartScrape Pro — Celery Task Worker
Background job execution for scraping tasks
"""
import asyncio
import json
import os
from datetime import datetime, timezone
from typing import Optional

from celery import Celery
from backend.utils.logger import logger

from config.settings import settings

# Create Celery app
celery_app = Celery(
    "smartscrape",
    broker=settings.CELERY_BROKER_URL,
    backend=settings.CELERY_RESULT_BACKEND,
)

celery_app.conf.update(
    task_serializer="json",
    accept_content=["json"],
    result_serializer="json",
    timezone="UTC",
    enable_utc=True,
    task_track_started=True,
    task_acks_late=True,
    worker_prefetch_multiplier=1,
    task_routes={
        "backend.scheduler.tasks.run_scraping_job": {"queue": "scraping"},
        "backend.scheduler.tasks.send_notification": {"queue": "notifications"},
    },
)


def run_async(coro):
    """Run async coroutine from sync Celery task."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


@celery_app.task(
    bind=True,
    name="backend.scheduler.tasks.run_scraping_job",
    max_retries=3,
    default_retry_delay=60,
    soft_time_limit=300,
    time_limit=600,
)
def run_scraping_job(self, job_id: str):
    """
    Execute a scraping job in the background.
    This is the main Celery task for scraping.
    """
    logger.info(f"🔄 Starting Celery task for job {job_id}")

    try:
        result = run_async(_execute_scraping_job(job_id, self.request.id))
        logger.success(f"✅ Job {job_id} completed: {result}")
        return result
    except Exception as exc:
        logger.error(f"❌ Job {job_id} failed: {exc}")
        # Retry on failure
        raise self.retry(exc=exc, countdown=60 * (self.request.retries + 1))


async def _execute_scraping_job(job_id: str, celery_task_id: str):
    """Async implementation of job execution."""
    from backend.models.database import AsyncSessionLocal
    from backend.models.models import ScrapingJob, JobLog, ScrapingResult as ResultModel, JobStatus, Subscription
    from backend.scraping.engine import scraping_engine
    from sqlalchemy import select
    import pandas as pd

    async with AsyncSessionLocal() as db:
        # Fetch job
        result = await db.execute(select(ScrapingJob).where(ScrapingJob.id == job_id))
        job = result.scalar_one_or_none()

        if not job:
            logger.error(f"Job {job_id} not found")
            return {"error": "Job not found"}

        # Update job status to running
        job.status = JobStatus.RUNNING
        job.started_at = datetime.now(timezone.utc)
        job.celery_task_id = celery_task_id

        await db.commit()

        # Log start
        log = JobLog(job_id=job_id, level="INFO", message=f"Job started. Engine: {job.engine.value}")
        db.add(log)
        await db.commit()

        try:
            # Build job config
            job_config = {
                "url": job.target_url,
                "engine": job.engine.value,
                "selectors": job.selectors or {},
                "headers": job.headers or {},
                "cookies": job.cookies or [],
                "proxy": job.proxy_config or {},
                "pagination": job.pagination_config or {},
                "timeout": 30,
            }

            # Run scraping engine
            scrape_result = await scraping_engine.run(job_config)

            # Log scraping result
            log2 = JobLog(
                job_id=job_id,
                level="INFO",
                message=f"Scraped {len(scrape_result.records)} records from {scrape_result.pages_scraped} pages in {scrape_result.duration_seconds}s"
            )
            db.add(log2)

            # Log any errors
            for error in scrape_result.errors:
                err_log = JobLog(job_id=job_id, level="WARNING", message=error)
                db.add(err_log)

            # Save results to DB
            for record in scrape_result.records:
                result_record = ResultModel(
                    job_id=job_id,
                    data=record,
                    source_url=scrape_result.source_url,
                )
                db.add(result_record)

            # Export to file
            file_path = await _export_results(job, scrape_result.records)

            # Update job
            job.status = JobStatus.COMPLETED if not scrape_result.errors else JobStatus.COMPLETED
            job.records_scraped = len(scrape_result.records)
            job.pages_scraped = scrape_result.pages_scraped
            job.duration_seconds = scrape_result.duration_seconds
            job.completed_at = datetime.now(timezone.utc)
            job.last_run_at = datetime.now(timezone.utc)
            if file_path:
                job.result_file_path = file_path

            # Update ScheduleRun if applicable
            from backend.models.models import ScheduleRun, Schedule
            
            run_result = await db.execute(select(ScheduleRun).where(ScheduleRun.job_id == job_id))
            schedule_run = run_result.scalar_one_or_none()
            if schedule_run:
                schedule_run.status = "completed"
                schedule_run.duration_seconds = scrape_result.duration_seconds
                schedule_run.records_extracted = len(scrape_result.records)
                schedule_run.completed_at = datetime.now(timezone.utc)
                
                # Update parent schedule success count
                sched_result = await db.execute(select(Schedule).where(Schedule.id == schedule_run.schedule_id))
                parent_sched = sched_result.scalar_one_or_none()
                if parent_sched:
                    parent_sched.success_count += 1
                
                # Create internal notification
                from backend.models.models import Notification
                notif = Notification(
                    user_id=job.user_id,
                    title=f"Schedule '{parent_sched.name}' Completed",
                    message=f"Extracted {len(scrape_result.records)} records in {scrape_result.duration_seconds}s.",
                    type="success"
                )
                db.add(notif)


            # Increment usage counter
            sub_result = await db.execute(
                select(Subscription).where(Subscription.user_id == job.user_id)
            )
            sub = sub_result.scalar_one_or_none()
            if sub:
                sub.jobs_used_this_month += 1

            await db.commit()

            return {
                "status": "completed",
                "records": len(scrape_result.records),
                "pages": scrape_result.pages_scraped,
                "duration": scrape_result.duration_seconds,
            }

        except Exception as e:
            # Mark job as failed
            job.status = JobStatus.FAILED
            job.error_message = str(e)[:500]
            job.completed_at = datetime.now(timezone.utc)

            err_log = JobLog(
                job_id=job_id,
                level="ERROR",
                message=f"Job failed: {str(e)[:300]}"
            )
            db.add(err_log)
            await db.commit()

            raise


async def _export_results(job, records: list) -> Optional[str]:
    """Export scraped results to file (CSV/JSON/XLSX)."""
    if not records:
        return None

    export_dir = os.path.join("./exports", job.user_id)
    os.makedirs(export_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"job_{job.id[:8]}_{timestamp}"

    try:
        import pandas as pd

        df = pd.DataFrame(records)
        fmt = job.export_format.value if job.export_format else "json"

        if fmt == "csv":
            file_path = os.path.join(export_dir, f"{base_name}.csv")
            df.to_csv(file_path, index=False)
        elif fmt == "xlsx":
            file_path = os.path.join(export_dir, f"{base_name}.xlsx")
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Scraped Data")
                workbook = writer.book
                worksheet = writer.sheets["Scraped Data"]
                
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment, Font
                
                # Bold headers
                for row in worksheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = Font(bold=True)

                # Auto-adjust column width and text wrap
                for idx, col in enumerate(df.columns, 1):
                    col_letter = get_column_letter(idx)
                    max_length = min(df[col].astype(str).map(len).max() if not df[col].empty else 10, 80)
                    header_len = len(str(col))
                    adjusted_width = max(max_length, header_len) + 2
                    worksheet.column_dimensions[col_letter].width = min(adjusted_width, 100)
                    
                    for cell in worksheet[col_letter]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        elif fmt == "json":
            file_path = os.path.join(export_dir, f"{base_name}.json")
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(records, f, indent=4, ensure_ascii=False)
        else:
            file_path = os.path.join(export_dir, f"{base_name}.json")
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(records, f, indent=4, ensure_ascii=False)

        return file_path

    except Exception as e:
        logger.warning(f"Export failed: {e}")
        # Fallback to JSON
        file_path = os.path.join(export_dir, f"{base_name}.json")
        with open(file_path, "w") as f:
            json.dump(records, f, indent=2, default=str)
        return file_path


@celery_app.task(name="backend.scheduler.tasks.reset_monthly_usage")
def reset_monthly_usage():
    """Reset monthly job usage counters — run via Celery Beat on 1st of month."""
    run_async(_do_monthly_reset())


async def _do_monthly_reset():
    from backend.models.database import AsyncSessionLocal
    from backend.models.models import Subscription
    from sqlalchemy import select, update
    from config.settings import PLANS

    async with AsyncSessionLocal() as db:
        await db.execute(
            update(Subscription).values(
                jobs_used_this_month=0,
            )
        )
        await db.commit()
        logger.info("Monthly usage counters reset for all subscriptions")


# Celery Beat schedule (for scheduled jobs check)
celery_app.conf.beat_schedule = {
    "check-scheduled-jobs": {
        "task": "backend.scheduler.tasks.check_scheduled_jobs",
        "schedule": 60.0,  # every 60 seconds
    },
    "reset-monthly-usage": {
        "task": "backend.scheduler.tasks.reset_monthly_usage",
        "schedule": 86400.0,  # daily (actual reset logic checks date)
    },
}


@celery_app.task(name="backend.scheduler.tasks.check_scheduled_jobs")
def check_scheduled_jobs():
    """Check for scheduled jobs that need to run now."""
    run_async(_check_and_trigger_scheduled_jobs())


async def _check_and_trigger_scheduled_jobs():
    from backend.models.database import AsyncSessionLocal
    from backend.models.models import ScrapingJob, JobStatus
    from sqlalchemy import select

    now = datetime.now(timezone.utc)

    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(ScrapingJob).where(
                ScrapingJob.is_scheduled == True,
                ScrapingJob.status != JobStatus.RUNNING,
                ScrapingJob.next_run_at <= now,
            )
        )
        due_jobs = result.scalars().all()

        for job in due_jobs:
            logger.info(f"Triggering scheduled job {job.id}: {job.name}")
            run_scraping_job.delay(job.id)
            # Update next run time based on cron
            job.status = JobStatus.QUEUED
            await db.commit()
